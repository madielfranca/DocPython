import pandas as pd
from openpyxl import load_workbook

df_carga = pd.read_excel('C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/PROJETO CARGA.xlsm', sheet_name='GL_SEGMENT_VALUES_INTERFACE')
df_hierarquia = pd.read_excel('C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/PP.OO.9.100 - Hierarquia de Projetos.xlsm', sheet_name='GL_SEGMENT_HIER_INTERFACE', header=0)

for row in df_carga.index:
    #Verifica se o nome da coluna é 'Unnamed'
    if 'Unnamed: 0' in df_carga.columns:
        valores_projeto_carga = df_carga['Unnamed: 1'][row]

    else:
        valores_projeto_carga = df_carga['B'][row]
    valores_projeto_carga = str(valores_projeto_carga)
    primeira_letra = valores_projeto_carga[0]

    quantidade_caracteres = len(valores_projeto_carga)
    numeros_aceitos = {10, 13, 16, 12}
    if quantidade_caracteres in numeros_aceitos:
        if primeira_letra.startswith(('P', 'E', 'I','C')):
            minha_lista = []
            for row in df_hierarquia.index:
                if row in df_hierarquia['Unnamed: 33'].index:
                    #Verifica se o nome da coluna é 'Unnamed'
                    if 'Unnamed: 0' in df_hierarquia.columns:
                        parent1Hierarquia = df_hierarquia['Unnamed: 33'][row]
                        setCode = df_hierarquia['Unnamed: 0'][row]
                        treeCode = df_hierarquia['Unnamed: 1'][row]
                        treeCodeVersion = df_hierarquia['Unnamed: 2'][row]
                        treeCodeVersionDate = df_hierarquia['Unnamed: 3'][row]
                    else:
                        parent1Hierarquia = df_hierarquia['AJ'][row]

                    parent1Hierarquia = str(parent1Hierarquia)
                    qtCaracteresParent1 = valores_projeto_carga[:-3]

                    if primeira_letra.startswith(('C')):
                        qtCaracteresParent1 = valores_projeto_carga[:-4]

                    # Verificando se o parent 3 existe
                    if parent1Hierarquia == qtCaracteresParent1:

                        # Verificando se a string começa com as letras P, E, I ou O
                        
                            quantidade_caracteres = len(valores_projeto_carga)
                            quantidade_caracteresP1 = len(qtCaracteresParent1)


                            # quantidade_caracteres
                            if quantidade_caracteres == 16:
                                # Define o índice onde deseja inserir a linha em branco
                                nova_linha = pd.DataFrame([[None] * len(df_hierarquia.columns)], columns=df_hierarquia.columns)

                                row = row+2
                                indice_insercao = row
                                ultimoItem = None
                                contador = 0
                                for i in range(row, len(df_hierarquia)):  
                                    if primeira_letra.startswith(('C')): 
                                        ColValue = df_hierarquia.at[i, 'Unnamed: 35']
                                    else:
                                        ColValue = df_hierarquia.at[i-1, 'Unnamed: 35']
                                    minha_lista.append(ColValue)

                                    if not isinstance(ColValue, str) or ColValue == "nan":
        
                                        if  contador > 0: 
                                            lista_sem_ultima = minha_lista[:-1]
                                            ultimo_valor = lista_sem_ultima[-1]

                                            # Encontre o índice do valor na coluna ''Unnamed: 35''
                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 35'] == ultimo_valor].tolist()
                                            
                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]+1], nova_linha, df_hierarquia.iloc[indice[0]+1:]], ignore_index=True)
                                            
                                            df_hierarquia.at[indice[0]+1, 'Unnamed: 35'] = valores_projeto_carga
                                            # df_hierarquia.at[indice[0]+1, 'Unnamed: 0'] = setCode
                                            # df_hierarquia.at[indice[0]+1, 'Unnamed: 1'] = treeCode
                                            # df_hierarquia.at[indice[0]+1, 'Unnamed: 2'] = treeCodeVersion
                                            # df_hierarquia.at[indice[0]+1, 'Unnamed: 3'] = treeCodeVersionDate

                                            break
                                        else:
                                            print('Nenhum filho') 

                                            break
                                    else:
                                        contador += 1
        
                                        ultimos_tres_caracteres_carga = valores_projeto_carga[-4:]
                                        ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)
                                    
                                        ultimos_tres_caracteres_parent1Hierarquia = parent1Hierarquia[-4:]
                                        ultimos_tres_caracteres_parent1Hierarquia = int(ultimos_tres_caracteres_parent1Hierarquia)

                                        ultimos_tres_caracteres_ColValue = ColValue[-4:]
                                        ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)


                                    if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                        print("O projeto duplicado.", valores_projeto_carga)
                                        break

                                    if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue:
                                        print("Projeto Incluido com sucesso.", valores_projeto_carga)

                                    # Encontre o índice do valor na coluna ''Unnamed: 35''
                                        indice = df_hierarquia.index[df_hierarquia['Unnamed: 35'] == ColValue].tolist()
                                        
                                        df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                        
                                        df_hierarquia.at[indice[0], 'Unnamed: 35'] = valores_projeto_carga
                                        # df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                        # df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                        # df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                        # df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate

                                        break



        else:
            print("O projeto" ,valores_projeto_carga, "não posui quantidade de letras validas.")
    else:                    
        if valores_projeto_carga != "nan" and valores_projeto_carga != "*Value":
            print("O projeto", valores_projeto_carga, " não possui letras validas, quantidade de letras .",quantidade_caracteres) 


# Caminho para o arquivo Excel existente
arquivo_excel = 'PP.OO.9.100 - Hierarquia de Projetos.xlsm'

# Carrega o DataFrame novo com os dados que deseja atualizar
df_novo = df_hierarquia

df_novo = df_novo.iloc[3:].reset_index(drop=True)

# Carrega o workbook existente
book = load_workbook(arquivo_excel, keep_vba=True)

# Seleciona a planilha a ser atualizada
sheet_name = 'GL_SEGMENT_HIER_INTERFACE'
if sheet_name not in book.sheetnames:
    raise ValueError(f"A planilha {sheet_name} não existe no arquivo Excel")

sheet = book[sheet_name]

# A linha onde os dados novos devem começar (abaixo da linha 4)
start_row = 5

# Atualiza as células abaixo da linha 4 com os dados do DataFrame
for row_index, row in df_novo.iterrows():
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=start_row + row_index, column=col_index + 1)
        cell.value = value

# Salva o workbook atualizado
book.save(arquivo_excel)


            

       