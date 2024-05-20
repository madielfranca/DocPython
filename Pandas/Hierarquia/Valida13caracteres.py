import pandas as pd
import math
from openpyxl import load_workbook

df_carga = pd.read_excel('C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/PROJETO CARGA.xlsm', sheet_name='GL_SEGMENT_VALUES_INTERFACE')
df_hierarquia = pd.read_excel('C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/PP.OO.9.100 - Hierarquia de Projetos.xlsm', sheet_name='GL_SEGMENT_HIER_INTERFACE', header=0)

for row in df_carga.index:
    #Verifica se o nome da coluna é 'Unnamed'
    if 'Unnamed: 0' in df_carga.columns:
        valores_projeto_carga = df_carga['Unnamed: 1'][row]

    else:
        valores_projeto_carga = df_carga['B'][row]
    valores_projeto_carga = str(valores_projeto_carga)
    primeira_letra = valores_projeto_carga[0]

    quantidade_caracteres = len(valores_projeto_carga)
    numeros_aceitos = {10, 13, 16, 12}
    if quantidade_caracteres in numeros_aceitos:
        if primeira_letra.startswith(('P', 'E', 'I','C')):
            minha_lista = []
            validaColValue = 0
            for row in df_hierarquia.index:
                if row in df_hierarquia['Unnamed: 32'].index:
                    #Verifica se o nome da coluna é 'Unnamed'
                    if 'Unnamed: 0' in df_hierarquia.columns:
                        parent1Hierarquia = df_hierarquia['Unnamed: 32'][row]
                        setCode = df_hierarquia['Unnamed: 0'][row]
                        treeCode = df_hierarquia['Unnamed: 1'][row]
                        treeCodeVersion = df_hierarquia['Unnamed: 2'][row]
                        treeCodeVersionDate = df_hierarquia['Unnamed: 3'][row]
                    else:
                        parent1Hierarquia = df_hierarquia['AJ'][row]
                 
                    parent1Hierarquia = str(parent1Hierarquia)
                    qtCaracteresParent1 = valores_projeto_carga[:-3]

                    if primeira_letra.startswith(('C')):
                        qtCaracteresParent1 = valores_projeto_carga[:-3]

                    # Verificando se o parent 3 existe
                    if parent1Hierarquia == qtCaracteresParent1:
  
                        # Verificando se a string começa com as letras P, E, I ou O
                        
                        quantidade_caracteres = len(valores_projeto_carga)
                        quantidade_caracteresP1 = len(qtCaracteresParent1)

                        # quantidade_caracteres
                        if quantidade_caracteres == 13:
                            # Define o índice onde deseja inserir a linha em branco
                            nova_linha = pd.DataFrame([[None] * len(df_hierarquia.columns)], columns=df_hierarquia.columns)

                            row = row+2
                            indice_insercao = row
                            ultimoItem = None
                            contador = 0
       
                            for i in range(row, len(df_hierarquia)):  
                                if primeira_letra.startswith(('C')): 
                                    ColValue = df_hierarquia.at[i-1, 'Unnamed: 33']
                                else:
                                    ColValue = df_hierarquia.at[i-1, 'Unnamed: 33']
                                minha_lista.append(ColValue)

                                if not isinstance(ColValue, str) or ColValue == "nan":
                                    
                                    if  contador > 0: 
                                        print('execao') 
  
                                        lista_sem_ultima = minha_lista[:-1]
                                        ultimo_valor = lista_sem_ultima[-1]

                                        # Encontre o índice do valor na coluna ''Unnamed: 35''
                                        indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ultimo_valor].tolist()
                                        
                                        df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]+1], nova_linha, df_hierarquia.iloc[indice[0]+1:]], ignore_index=True)
                                        
                                        if primeira_letra.startswith(('C')): 
                                            print(ultimos_tres_caracteres_carga) 
                                            print(ultimos_tres_caracteres_ColValue) 
                                            if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                print("O projeto duplicado.", valores_projeto_carga)
                                                break
                                            df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                        else:
                                            df_hierarquia.at[indice[0], 'Unnamed: 33'] = valores_projeto_carga
                                        # df_hierarquia.at[indice[0]+1, 'Unnamed: 0'] = setCode
                                        # df_hierarquia.at[indice[0]+1, 'Unnamed: 1'] = treeCode
                                        # df_hierarquia.at[indice[0]+1, 'Unnamed: 2'] = treeCodeVersion
                                        # df_hierarquia.at[indice[0]+1, 'Unnamed: 3'] = treeCodeVersionDate

                                        break
                                    else:
                                        print('Nenhum filho') 

                                        break
                                else:
                                    contador += 1

                                    if primeira_letra.startswith(('C')) : 
                                        for c in range(row, len(df_hierarquia)): 
                                            ColValue = df_hierarquia.at[c, 'Unnamed: 34']
                                            if not (ColValue is None or (isinstance(ColValue, float) and math.isnan(ColValue))):
                                                # Se my_var não for None, execute o código aqui
                                                # print("my_var tem um valor válido:", ColValue)
                                                ultimos_tres_caracteres_carga = valores_projeto_carga[-5:]
                                                ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)

                                                ultimos_tres_caracteres_ColValue = ColValue[-5:]
                                                ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)

                                                comparaCaractereP1 = valores_projeto_carga[:-3]
                                                comparaCaractereP2 = ColValue[:-3]
                                                # print(validaColValue) 
                                                # print(ultimos_tres_caracteres_ColValue) 
                                                # print(ultimos_tres_caracteres_carga) 
                                                # print(ultimos_tres_caracteres_ColValue) 
                                                # print('---------') 
                                                if validaColValue < ultimos_tres_caracteres_ColValue and ultimos_tres_caracteres_carga > ultimos_tres_caracteres_ColValue:
                                                    validaColValue = ultimos_tres_caracteres_ColValue 
                                                    ColValueP2 = ColValue 
                 
                                                # print(comparaCaractereP1) 
                                                # print(comparaCaractereP2)
                                                # print(ultimos_tres_caracteres_carga) 
                                                # print(ultimos_tres_caracteres_ColValue) 
                                                if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                    print("O projeto duplicado.", valores_projeto_carga)
                                                    
                                                    break
                                         
                                                if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 != comparaCaractereP2:
                                                    # print('passou') 
                                                    # print(comparaCaractereP1) 
                                                    # print(comparaCaractereP2) 
                                                    # breakpoint()
                                                        # Encontre o índice do valor na coluna ''Unnamed: 35''
                                                    indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                    
                                                    df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-2], nova_linha, df_hierarquia.iloc[indice[0]-2:]], ignore_index=True)
                                                    if primeira_letra.startswith(('C')): 
                                                        df_hierarquia.at[indice[0]-2, 'Unnamed: 34'] = valores_projeto_carga
                                                    else:
                                                        df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate

                                                    break
                                                elif ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 == comparaCaractereP2:
                                                    indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                    # breakpoint()
                                                    # ColValueP2 caso nessecite inserir a linha abaixo do penultimo item
                                                    # print(ColValueP2) 
                                                    print(ultimos_tres_caracteres_carga)  
                                                    print(ultimos_tres_caracteres_ColValue)  
                                                    print('else') 
                                                    print(ColValueP2) 
                                                    print(comparaCaractereP1) 
                                                    print(comparaCaractereP2) 
                                                    df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                                    if primeira_letra.startswith(('C')): 
                                                        df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                        # df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                    break
                                        # breakpoint()
                                        break
                                    else:
                                        ultimos_tres_caracteres_carga = valores_projeto_carga[-4:]
                                        ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)

                                        ultimos_tres_caracteres_parent1Hierarquia = parent1Hierarquia[-4:]
                                        ultimos_tres_caracteres_parent1Hierarquia = int(ultimos_tres_caracteres_parent1Hierarquia)

                                        ultimos_tres_caracteres_ColValue = ColValue[-4:]
                                        ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)

                                        if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                            print("O projeto duplicado.", valores_projeto_carga)
                                            break

                                        if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue:
                                            print("Projeto Incluido com sucesso.", valores_projeto_carga)

                                            # Encontre o índice do valor na coluna ''Unnamed: 35''
                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ColValue].tolist()
                                            
                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                            if primeira_letra.startswith(('C')): 
                                                df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                            else:
                                                df_hierarquia.at[indice[0], 'Unnamed: 33'] = valores_projeto_carga
                                                # df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                # df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                # df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                # df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate

                                            break

        else:
            print("O projeto" ,valores_projeto_carga, "não posui quantidade de letras validas.")
    else:                    
        if valores_projeto_carga != "nan" and valores_projeto_carga != "*Value":
            print("O projeto", valores_projeto_carga, " não possui letras validas, quantidade de letras .",quantidade_caracteres) 

# df_hierarquia = df_hierarquia.drop([0, 1])


# Abre o arquivo Excel existente
arquivo_excel = 'PP.OO.9.100 - Hierarquia de Projetos.xlsm'

# Carrega o arquivo Excel existente
book = load_workbook(arquivo_excel)

# Carrega o DataFrame novo com os dados que deseja atualizar
df_novo = df_hierarquia

# Carrega o workbook existente
book = load_workbook(arquivo_excel)

# Adiciona ou atualiza a planilha
with pd.ExcelWriter(arquivo_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    writer.book = book
    
    # Especifica a planilha a ser atualizada
    df_novo.to_excel(writer, sheet_name='GL_SEGMENT_HIER_INTERFACE', index=False)

# Salva o workbook atualizado
book.save(arquivo_excel)

# # Selecionar apenas as colunas 'A', 'C', e 'D'
# selected_columns = ['Unnamed: 32', 'Unnamed: 33', 'Unnamed: 34', 'Unnamed: 35']
# df_selected = df_hierarquia[selected_columns]


# nome_arquivo = 'HierarquiaDeProjetos13.xlsx'
# df_selected.to_excel(nome_arquivo, index=False, header=None )



            

       