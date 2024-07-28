from globo_automacoes.decoradores import retry
from src.LogStatus import LogStatus
from openpyxl import load_workbook
import pandas as pd
import logging

log = logging.getLogger(__name__)

class Hierarquia12Caracteres:
    def __init__(self, projeto_carga_excel_path, projeto_carga_sheet_name, projeto_hierarquia_excel_path, projeto_hierarquia_sheet_name, filename) -> None:
        self.df_carga = pd.read_excel(projeto_carga_excel_path, sheet_name=projeto_carga_sheet_name, header=0)
        self.df_hierarquia = pd.read_excel(projeto_hierarquia_excel_path, sheet_name=projeto_hierarquia_sheet_name, header=0)
        self.filename = filename

    @retry(3, Exception)
    def validar_projetos_12_caracteres(self):
        log.info("Iniciando execução da task Hierarquia 12 caracteres")    

        df_carga = self.df_carga
        file_name_caraga = self.filename
        df_hierarquia = pd.read_excel('C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/files/Hierarquia/PP.OO.9.100 - Hierarquia de Projetos.xlsm', sheet_name='GL_SEGMENT_HIER_INTERFACE', header=0)
        values_to_add  = []
        for row in df_carga.index:
            #Verifica se o nome da coluna é 'Unnamed'
            if 'Unnamed: 0' in df_carga.columns:
                valores_projeto_carga = df_carga['Unnamed: 1'][row]

            else:
                valores_projeto_carga = df_carga['B'][row]
            valores_projeto_carga = str(valores_projeto_carga)
            primeira_letra = valores_projeto_carga[0]

            quantidade_caracteres = len(valores_projeto_carga)
            nova_linha = pd.DataFrame([[None] * len(df_hierarquia.columns)], columns=df_hierarquia.columns)
            numeros_aceitos = {12}
            if quantidade_caracteres in numeros_aceitos:
                if primeira_letra.startswith(('P', 'E', 'I','C')):
                    minha_lista = []
                    contador = 0
                    for row in df_hierarquia.index:
                        if row in df_hierarquia['Unnamed: 31'].index:
                            #Verifica se o nome da coluna é 'Unnamed'
                            if 'Unnamed: 0' in df_hierarquia.columns:
                                if primeira_letra.startswith(('C')):
                                    parent3Hierarquia = df_hierarquia['Unnamed: 32'][row]
                                else:
                                    parent3Hierarquia = df_hierarquia['Unnamed: 33'][row]

                                setCode = df_hierarquia['Unnamed: 0'][row]
                                treeCode = df_hierarquia['Unnamed: 1'][row]
                                treeCodeVersion = df_hierarquia['Unnamed: 2'][row]
                                treeCodeVersionDate = df_hierarquia['Unnamed: 3'][row]
                            else:
                                parent3Hierarquia = df_hierarquia['AJ'][row]

                            parent3Hierarquia = str(parent3Hierarquia)
                            if primeira_letra.startswith(('C')):
                                parent3Hierarquia = parent3Hierarquia
                                qtCaracteresParent3 = valores_projeto_carga[:-2]
                            else:
                                parent3Hierarquia = parent3Hierarquia[:-2]
                                qtCaracteresParent3 = valores_projeto_carga
                            conCaracteresParent3 = valores_projeto_carga[-3:]
 
                            if parent3Hierarquia == qtCaracteresParent3:
                                if contador < 1:
                                    for c in range(row, len(df_hierarquia)):
                                        ColValue = df_hierarquia.at[c, 'Unnamed: 32']
                                        ColValue = str(ColValue)
                                        conCaracteresParent3Hierarquia = ColValue[-3:]
                                        parent4 = ColValue[:-2]
                                        if ColValue != "nan":
                                            # Verificando se o parent 3 existe
                                            if ColValue == valores_projeto_carga:
                                                print("Projeto Incluido com sucesso.", valores_projeto_carga)
                                                print('1')
                                                values_to_add.append("O projeto duplicado."+ valores_projeto_carga)
                                                break
                                
                                                
                                            elif conCaracteresParent3 < conCaracteresParent3Hierarquia and parent3Hierarquia == parent4:
                                                # Encontre o índice do valor na coluna ''Unnamed: 32'
                                                indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ColValue].tolist()                                                
                                                df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)

                                                df_hierarquia.at[indice[0], 'Unnamed: 33'] = valores_projeto_carga
                                                df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                print("Projeto Incluido com sucesso.", valores_projeto_carga)
                                                print('2')
                                                values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)
                                                contador += 1

                                                break

                                        
                                            elif qtCaracteresParent3 != parent4:
                                                # Encontre o índice do valor na coluna ''Unnamed: 35''
                                                indice = df_hierarquia.index[df_hierarquia['Unnamed: 32'] == ColValue].tolist()
                                                print(ColValue)
                                                print(indice)
                                                # breakpoint()
                                                df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]+1], nova_linha, df_hierarquia.iloc[indice[0]+1:]], ignore_index=True)
                                                
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 33'] = valores_projeto_carga
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 0'] = setCode
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 1'] = treeCode
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 2'] = treeCodeVersion
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 3'] = treeCodeVersionDate
                                                values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)
                                                print('3')
                                                print("Projeto Incluido com sucesso.", valores_projeto_carga)
                                                contador += 1

                                                break
                                    break     

        LogStatus_obj=LogStatus(values_to_add, file_name_caraga)
        LogStatus_obj.logar()

        # Caminho para o arquivo Excel existente
        arquivo_excel = 'C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/files/Hierarquia/PP.OO.9.100 - Hierarquia de Projetos.xlsm'

        # Carrega o DataFrame novo com os dados que deseja atualizar
        df_novo = df_hierarquia

        df_novo = df_novo.iloc[3:].reset_index(drop=True)

        # Carrega o workbook existente
        book = load_workbook(arquivo_excel, keep_vba=True)

        # Seleciona a planilha a ser atualizada
        sheet_name = 'GL_SEGMENT_HIER_INTERFACE'
        if sheet_name not in book.sheetnames:
            raise ValueError(f"A planilha {sheet_name} não existe no arquivo Excel")

        sheet = book[sheet_name]

        # A linha onde os dados novos devem começar (abaixo da linha 4)
        start_row = 5

        # Atualiza as células abaixo da linha 4 com os dados do DataFrame
        for row_index, row in df_novo.iterrows():
            for col_index, value in enumerate(row):
                cell = sheet.cell(row=start_row + row_index, column=col_index + 1)
                cell.value = value

        # Salva o workbook atualizado
        book.save(arquivo_excel)
        log.info("Hierarquia 12 caracteres executado com sucesso")


            

       