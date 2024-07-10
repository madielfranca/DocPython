from globo_automacoes.decoradores import retry
from openpyxl import load_workbook
from src.LogStatus import LogStatus
import pandas as pd
import math

class Hierarquia13Caracteres:
    def __init__(self, projeto_carga_excel_path, projeto_carga_sheet_name, projeto_hierarquia_excel_path, projeto_hierarquia_sheet_name, filename) -> None:
        self.df_carga = pd.read_excel(projeto_carga_excel_path, sheet_name=projeto_carga_sheet_name, header=0)
        self.df_hierarquia = pd.read_excel(projeto_hierarquia_excel_path, sheet_name=projeto_hierarquia_sheet_name, header=0)
        self.filename = filename

    @retry(3, Exception)
    def validar_projetos_13_caracteres(self):
        # print(self.df_carga)

        df_carga = self.df_carga
        df_hierarquia = self.df_hierarquia
        values_to_add  = []
        file_name_caraga = self.filename
        for row in df_carga.index:
            #Verifica se o nome da coluna é 'Unnamed'
            if 'Unnamed: 0' in df_carga.columns:
                valores_projeto_carga = df_carga['Unnamed: 1'][row]

            else:
                valores_projeto_carga = df_carga['B'][row]
            valores_projeto_carga = str(valores_projeto_carga)
            primeira_letra = valores_projeto_carga[0]

            quantidade_caracteres = len(valores_projeto_carga)
            numeros_aceitos = {10, 13, 16, 12}
            if quantidade_caracteres in numeros_aceitos:
                if primeira_letra.startswith(('P', 'E', 'I','C')):
                    minha_lista = []
                    validaColValue = 0
                    for row in df_hierarquia.index:
                        if row in df_hierarquia['Unnamed: 32'].index:
                            #Verifica se o nome da coluna é 'Unnamed'
                            if 'Unnamed: 0' in df_hierarquia.columns:
                                parent1Hierarquia = df_hierarquia['Unnamed: 32'][row]
                                setCode = df_hierarquia['Unnamed: 0'][row]
                                treeCode = df_hierarquia['Unnamed: 1'][row]
                                treeCodeVersion = df_hierarquia['Unnamed: 2'][row]
                                treeCodeVersionDate = df_hierarquia['Unnamed: 3'][row]
                            else:
                                parent1Hierarquia = df_hierarquia['AJ'][row]
                        
                            parent1Hierarquia = str(parent1Hierarquia)
                            qtCaracteresParent1 = valores_projeto_carga[:-3]

                            if primeira_letra.startswith(('C')):
                                qtCaracteresParent1 = valores_projeto_carga[:-3]

                            # Verificando se o parent 3 existe
                            if parent1Hierarquia == qtCaracteresParent1:
        
                                # Verificando se a string começa com as letras P, E, I ou O
                                
                                quantidade_caracteres = len(valores_projeto_carga)
                                quantidade_caracteresP1 = len(qtCaracteresParent1)

                                # quantidade_caracteres
                                if quantidade_caracteres == 13:
                                    # Define o índice onde deseja inserir a linha em branco
                                    nova_linha = pd.DataFrame([[None] * len(df_hierarquia.columns)], columns=df_hierarquia.columns)

                                    row = row+2
                                    indice_insercao = row
                                    ultimoItem = None
                                    contador = 0
                                    ColValueP2 = 0
            
                                    for i in range(row, len(df_hierarquia)):  
                                               
                                        ColValue = df_hierarquia.at[i-1, 'Unnamed: 33']
                                        minha_lista.append(ColValue)

                                        if not isinstance(ColValue, str) or ColValue == "nan":                                                    
                                            if  contador > 0: 
        
                                                lista_sem_ultima = minha_lista[:-1]
                                                ultimo_valor = lista_sem_ultima[-1]

                                                # Encontre o índice do valor na coluna ''Unnamed: 35''
                                                indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ultimo_valor].tolist()
                                                df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                                
                                                if primeira_letra.startswith(('C')): 
                                                    print(ultimos_tres_caracteres_carga) 
                                                    print(ultimos_tres_caracteres_ColValue) 
                                                    if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                        print("O projeto duplicado.", valores_projeto_carga)
                                                        values_to_add.append("O projeto duplicado."+ valores_projeto_carga)
                                                        break
                                                    teste = int(9111)
                                                    df_hierarquia.at[indice[0], 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                else:
                                                    if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                        print("O projeto duplicado.", valores_projeto_carga)
                                                        values_to_add.append("O projeto duplicado."+ valores_projeto_carga)
                                                        break
                                                    df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                                df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)

                                                break
                                            else:
                                                teste = int(555)
                                                ColValue = df_hierarquia.at[i-2, 'Unnamed: 32']
                                                indice = df_hierarquia.index[df_hierarquia['Unnamed: 32'] == ColValue].tolist()
                                                                                               
                                                df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]+1], nova_linha, df_hierarquia.iloc[indice[0]+1:]], ignore_index=True)

                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 33'] = valores_projeto_carga
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 0'] = setCode
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 1'] = treeCode
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 2'] = treeCodeVersion
                                                df_hierarquia.at[indice[0]+1, 'Unnamed: 3'] = treeCodeVersionDate
                                            
                                                values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)

                                                break
                                        else:
                                            contador += 1

                                            if primeira_letra.startswith(('C')) : 
                                                for c in range(row, len(df_hierarquia)): 
                                                    ColValue = df_hierarquia.at[c-2, 'Unnamed: 33']
                                                    if not (ColValue is None or (isinstance(ColValue, float) and math.isnan(ColValue))):
                                                        # Se my_var não for None, execute o código aqui
                                                        ultimos_tres_caracteres_carga = valores_projeto_carga[-5:]
                                                        ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)

                                                        ultimos_tres_caracteres_ColValue = ColValue[-5:]
                                                        ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)

                                                        comparaCaractereP1 = valores_projeto_carga[:-3]
                                                        comparaCaractereP2 = ColValue[:-3]
                                                        if validaColValue < ultimos_tres_caracteres_ColValue and ultimos_tres_caracteres_carga > ultimos_tres_caracteres_ColValue:
                                                            validaColValue = ultimos_tres_caracteres_ColValue 
                                                            ColValueP2 = ColValue 

                                                        if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                            print("O projeto duplicado.", valores_projeto_carga)
                                                            values_to_add.append("O projeto duplicado."+ valores_projeto_carga)
                                                            
                                                            break
                                                
                                                        if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 != comparaCaractereP2:
                                                            values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)

                                                            # Encontre o índice do valor na coluna 'Unnamed: 35'
                                                            def validar_primeiras_letras(string, letras):
                                                                return string[:4] == letras

                                                            # Entrada do usuário
                                                            string_teste = valores_projeto_carga
                                                            letras_desejadas = "CFJI"
                                                            # Validação
                                                            if validar_primeiras_letras(string_teste, letras_desejadas):
                                                                for c in range(row, len(df_hierarquia)):
                                                                    ColValue = df_hierarquia.at[c, 'Unnamed: 33']
                                                                    # ColValue = df_hierarquia.at[row, 'Unnamed: 33']
                                                                    print(ColValue)
                                                                    print(row)
                                                                    print(c)
                                                                    # breakpoint()   
                                                                    if not (ColValue is None or (isinstance(ColValue, float) and math.isnan(ColValue))):
                                                                        print(row)
                                                                        print(ColValue)
                                                                        print(valores_projeto_carga)
                                                                        ultimos_tres_caracteres_carga = valores_projeto_carga[-6:]
                                                                        ultimos_tres_caracteres_carga = ultimos_tres_caracteres_carga[:-1]
                                                                        ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)
                                                                        ultimos_tres_caracteres_ColValue = ColValue[-5:]

                                                                        ultimos_tres_caracteres_ColValue = ColValue[-5:]
                                                                        ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)

                                                                        if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue:
                                                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ColValue].tolist()
                                                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-1], nova_linha, df_hierarquia.iloc[indice[0]-1:]], ignore_index=True)
                                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 34'] = valores_projeto_carga
                                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 0'] = setCode
                                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 1'] = treeCode
                                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 2'] = treeCodeVersion
                                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 3'] = treeCodeVersionDate
                                                                            break
                                                                        else:
                                                                            print('teste')
                                                                            # ColValue = df_hierarquia.at[c, 'Unnamed: 34']
                                                                            # indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                                            # df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-2], nova_linha, df_hierarquia.iloc[indice[0]-2:]], ignore_index=True)
                                                                            # teste = int(3)
                                                                            # df_hierarquia.at[indice[0], 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                                            # df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                                            # df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                                            # df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                                            # df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                                        
                                                                        # if primeira_letra.startswith(('CCFJI')): 
                                                                        #     teste = int(1999)
                                                                        #     df_hierarquia.at[indice[0]-2, 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                                        #     df_hierarquia.at[indice[0]-2, 'Unnamed: 0'] = setCode
                                                                        #     df_hierarquia.at[indice[0]-2, 'Unnamed: 1'] = treeCode
                                                                        #     df_hierarquia.at[indice[0]-2, 'Unnamed: 2'] = treeCodeVersion
                                                                        #     df_hierarquia.at[indice[0]-2, 'Unnamed: 3'] = treeCodeVersionDate
                                                                        # else:
                                                                        #     ColValue = df_hierarquia.at[c, 'Unnamed: 34']
                                                                        #     indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                                        #     df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-2], nova_linha, df_hierarquia.iloc[indice[0]-2:]], ignore_index=True)
                                                                        #     teste = int(3)
                                                                        #     df_hierarquia.at[indice[0], 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                                        #     df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                                        #     df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                                        #     df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                                        #     df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate

                                                                        # break
                                                                break
                                                            else:
                                                                ColValue = df_hierarquia.at[c, 'Unnamed: 34']
                                                                indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                                df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-2], nova_linha, df_hierarquia.iloc[indice[0]-2:]], ignore_index=True)

                                                                df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                                                df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                                df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                                df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                                df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                            
                                                            # if primeira_letra.startswith(('CCFJI')): 
                                                            #     teste = int(1999)
                                                            #     df_hierarquia.at[indice[0]-2, 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                            #     df_hierarquia.at[indice[0]-2, 'Unnamed: 0'] = setCode
                                                            #     df_hierarquia.at[indice[0]-2, 'Unnamed: 1'] = treeCode
                                                            #     df_hierarquia.at[indice[0]-2, 'Unnamed: 2'] = treeCodeVersion
                                                            #     df_hierarquia.at[indice[0]-2, 'Unnamed: 3'] = treeCodeVersionDate
                                                            # else:
                                                            #     ColValue = df_hierarquia.at[c, 'Unnamed: 34']
                                                            #     indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
                                                            #     df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-2], nova_linha, df_hierarquia.iloc[indice[0]-2:]], ignore_index=True)
                                                            #     teste = int(3)
                                                            #     df_hierarquia.at[indice[0], 'Unnamed: 34'] = f'{valores_projeto_carga}{teste}'
                                                            #     df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                            #     df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                            #     df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                            #     df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate

                                                                break
                                                        elif ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 == comparaCaractereP2:
                                                            values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)
                                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 34'] == ColValue].tolist()
 
                                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                                            if primeira_letra.startswith(('C')): 
                                                                df_hierarquia.at[indice[0], 'Unnamed: 34'] = valores_projeto_carga
                                                                df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                                df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                                df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                                df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                            break
                                                break
                                            else:
                                                for c in range(row, len(df_hierarquia)): 
                                                    ColValue = df_hierarquia.at[c-1, 'Unnamed: 33']
                                                  
                                                    if not (ColValue is None or (isinstance(ColValue, float) and math.isnan(ColValue))):
                                                        # Se my_var não for None, execute o código aqui
                                                        ultimos_tres_caracteres_carga = valores_projeto_carga[-5:]
                                                        ultimos_tres_caracteres_carga = int(ultimos_tres_caracteres_carga)

                                                        ultimos_tres_caracteres_ColValue = ColValue[-5:]
                                                        ultimos_tres_caracteres_ColValue = int(ultimos_tres_caracteres_ColValue)

                                                        comparaCaractereP1 = valores_projeto_carga[:-3]
                                                        comparaCaractereP2 = ColValue[:-3]

                                                        #Guarda o penultimo valor da hierarquia 
                                                        if validaColValue < ultimos_tres_caracteres_ColValue and ultimos_tres_caracteres_carga > ultimos_tres_caracteres_ColValue:
                                                            validaColValue = ultimos_tres_caracteres_ColValue 
                                                            ColValueP2 = ColValue 

                                                        if ultimos_tres_caracteres_carga == ultimos_tres_caracteres_ColValue:
                                                            print("O projeto duplicado.", valores_projeto_carga)
                                                            
                                                            break
                                                
                                                        if ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 != comparaCaractereP2:
                                                            # Encontre o índice do valor na coluna ''Unnamed: 35''
                                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ColValue].tolist()
                                               
                                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]-1], nova_linha, df_hierarquia.iloc[indice[0]-1:]], ignore_index=True)
                                                            values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)
                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 33'] = valores_projeto_carga
                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 0'] = setCode
                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 1'] = treeCode
                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 2'] = treeCodeVersion
                                                            df_hierarquia.at[indice[0]-1, 'Unnamed: 3'] = treeCodeVersionDate

                                                            break
                                                        elif ultimos_tres_caracteres_carga < ultimos_tres_caracteres_ColValue and comparaCaractereP1 == comparaCaractereP2:
                                                            values_to_add.append("Projeto Incluido com sucesso."+ valores_projeto_carga)
                                                            indice = df_hierarquia.index[df_hierarquia['Unnamed: 33'] == ColValue].tolist()

                                                            df_hierarquia = pd.concat([df_hierarquia.iloc[:indice[0]], nova_linha, df_hierarquia.iloc[indice[0]:]], ignore_index=True)
                                                            df_hierarquia.at[indice[0], 'Unnamed: 33'] = valores_projeto_carga
                                                            df_hierarquia.at[indice[0], 'Unnamed: 0'] = setCode
                                                            df_hierarquia.at[indice[0], 'Unnamed: 1'] = treeCode
                                                            df_hierarquia.at[indice[0], 'Unnamed: 2'] = treeCodeVersion
                                                            df_hierarquia.at[indice[0], 'Unnamed: 3'] = treeCodeVersionDate
                                                            break
                                                break
                                    break

        LogStatus_obj=LogStatus(values_to_add, file_name_caraga)
        LogStatus_obj.logar()

        # Caminho para o arquivo Excel existente
        arquivo_excel = 'C:/Users/madis/Documents/DocPython/Pandas/Hierarquia/files/Hierarquia/PP.OO.9.100 - Hierarquia de Projetos.xlsm'

        # Carrega o DataFrame novo com os dados que deseja atualizar
        df_novo = df_hierarquia

        df_novo = df_novo.iloc[3:].reset_index(drop=True)

        # Carrega o workbook existente
        book = load_workbook(arquivo_excel, keep_vba=True)

        # Seleciona a planilha a ser atualizada
        sheet_name = 'GL_SEGMENT_HIER_INTERFACE'
        if sheet_name not in book.sheetnames:
            raise ValueError(f"A planilha {sheet_name} não existe no arquivo Excel")

        sheet = book[sheet_name]

        # A linha onde os dados novos devem começar (abaixo da linha 4)
        start_row = 5

        # Atualiza as células abaixo da linha 4 com os dados do DataFrame
        for row_index, row in df_novo.iterrows():
            for col_index, value in enumerate(row):
                cell = sheet.cell(row=start_row + row_index, column=col_index + 1)
                cell.value = value

        # Salva o workbook atualizado
        book.save(arquivo_excel)


            
            

       